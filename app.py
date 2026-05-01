"""
app.py — MailSort, головний модуль веб-застосунку.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "core"))

import functools
from flask import (Flask, render_template, redirect, url_for,
                   request, jsonify, flash, session, Response)

from config import SECRET_KEY, CATEGORIES as DEFAULT_CATEGORIES
from database import (
    init_db, get_emails_by_category, get_email_by_id,
    mark_as_read, get_stats, reclassify_email,
    search_emails, get_all_emails_for_export,
    get_setting, set_setting,
    get_spam_rules, add_spam_rule, delete_spam_rule,
    get_all_categories, get_category_names,
    add_category, delete_category,
    toggle_starred, get_starred_emails, get_starred_count,
    delete_email, delete_demo_emails, delete_all_emails,
)
from markupsafe import Markup, escape

def _lazy_classify(subject, body, sender=""):
    from classifier import classify
    return classify(subject, body, sender=sender)
import re as _re


def _render_body(text):
    """
    Перетворює plain-text або HTML тіло листа на безпечний HTML:
    - URL стають клікабельними посиланнями
    - подвійний \n стає абзацом, одинарний — <br>
    - якщо тіло вже HTML — прибираємо script/style і небезпечні атрибути
    """
    if not text:
        return Markup("")
    if "<html" in text.lower() or "<!doctype" in text.lower():
        clean = _re.sub(r'<(script|style)[^>]*>.*?</\1>', '', text,
                        flags=_re.IGNORECASE | _re.DOTALL)
        clean = _re.sub(r'\s+on\w+=["\'][^"\']*["\']', '', clean, flags=_re.IGNORECASE)
        return Markup(clean)
    safe = str(escape(text))
    url_pat = _re.compile(r'(https?://[^\s\)\]\'"&<>]+)', _re.IGNORECASE)
    safe = url_pat.sub(
        lambda m: f'<a href="{m.group(1)}" target="_blank" rel="noopener">{m.group(1)}</a>',
        safe
    )
    parts = _re.split(r'\n{2,}', safe)
    html  = "".join(f"<p>{p.replace(chr(10), '<br>')}</p>" for p in parts if p.strip())
    return Markup(html)

# ── Автосинхронізація ─────────────────────────────────────────────
import threading, time


_sync_thread   = None
_sync_stop     = threading.Event()
_sync_interval = 300   # секунд між перевірками (за замовчуванням 5 хв)
_last_sync_result = {"time": None, "fetched": 0, "status": "idle"}


# ── Стан ручної синхронізації (для прогрес-бару у UI) ─────────────
_manual_sync_lock = threading.Lock()
_manual_sync_state = {
    "phase":    "idle",
    "current":  0,
    "total":    0,
    "saved":    0,
    "skipped":  0,
    "errors":   0,
    "message":  "",
    "last_subject":  "",
    "last_category": "",
    "started_at":    None,
    "finished_at":   None,
}
_manual_sync_thread = None


def _progress_callback(payload: dict):
    with _manual_sync_lock:
        _manual_sync_state.update(payload)


def _run_manual_sync():
    try:
        from email_fetcher import fetch_and_classify
        with _manual_sync_lock:
            _manual_sync_state.update({
                "phase": "connecting", "current": 0, "total": 0,
                "saved": 0, "skipped": 0, "errors": 0,
                "message": "Запуск синхронізації…",
                "last_subject": "", "last_category": "",
                "started_at":  time.time(),
                "finished_at": None,
            })
        s = fetch_and_classify(progress_callback=_progress_callback)
        with _manual_sync_lock:
            if s.get("error_msg"):
                _manual_sync_state["phase"]   = "error"
                _manual_sync_state["message"] = s["error_msg"]
            else:
                _manual_sync_state["phase"]   = "done"
                _manual_sync_state["message"] = f"Готово: +{s.get('saved', 0)} нових"
            _manual_sync_state["finished_at"] = time.time()
        _last_sync_result.update({
            "time":    time.strftime("%H:%M:%S"),
            "fetched": s.get("fetched", 0),
            "status":  "error" if s.get("error_msg") else "ok",
        })
    except Exception as e:
        with _manual_sync_lock:
            _manual_sync_state["phase"]       = "error"
            _manual_sync_state["message"]     = str(e)
            _manual_sync_state["finished_at"] = time.time()


def _auto_sync_loop():
    """Фоновий потік — перевіряє пошту кожні N секунд."""
    while not _sync_stop.wait(_sync_interval):
        try:
            from email_fetcher import fetch_and_classify
            from database import get_setting
            if get_setting("email_user", ""):
                s = fetch_and_classify()
                _last_sync_result.update({
                    "time":    time.strftime("%H:%M:%S"),
                    "fetched": s.get("fetched", 0),
                    "status":  "error" if s.get("error_msg") else "ok",
                })
        except Exception:
            pass


def start_auto_sync(interval: int = 300):
    global _sync_thread, _sync_interval
    _sync_interval = interval
    _sync_stop.clear()
    if _sync_thread is None or not _sync_thread.is_alive():
        _sync_thread = threading.Thread(target=_auto_sync_loop, daemon=True)
        _sync_thread.start()


def stop_auto_sync():
    _sync_stop.set()

# ── Demo emails ──────────────────────────────────────────────────
DEMO_EMAILS = [
    {"uid":"demo_001","sender":"rector@university.edu.ua",
     "subject":"Запрошення на міжнародну наукову конференцію",
     "body":"Шановні колеги, запрошуємо вас взяти участь у Міжнародному симпозіумі з інформаційних технологій у Варшаві. Організаційний взнос покривається грантом.",
     "date":"2024-11-15T10:30:00"},
    {"uid":"demo_002","sender":"priyom@university.edu.ua",
     "subject":"Умови вступу та перелік документів для абітурієнтів",
     "body":"Шановні абітурієнти! Прийом документів розпочинається 1 липня. Для вступу необхідно надати: атестат, сертифікати ЗНО, заяву та фото. Мінімальний конкурсний бал — 150.",
     "date":"2024-11-14T09:00:00"},
    {"uid":"demo_003","sender":"hr@university.edu.ua",
     "subject":"Оголошення про вакансію доцента кафедри ІТ",
     "body":"Університет оголошує конкурс на заміщення вакантної посади доцента кафедри інформаційних технологій. Вимоги: науковий ступінь кандидата наук, досвід викладання від 3 років.",
     "date":"2024-11-13T14:20:00"},
    {"uid":"demo_004","sender":"dean@cs.university.edu.ua",
     "subject":"Розклад занять на зимовий семестр 2024/2025",
     "body":"Шановні студенти! Розклад занять на зимовий семестр опубліковано на порталі університету. Початок занять — 3 вересня.",
     "date":"2024-11-12T08:45:00"},
    {"uid":"demo_005","sender":"science@university.edu.ua",
     "subject":"Результати конкурсу наукових грантів молодих вчених",
     "body":"Оголошуємо результати конкурсу на отримання наукових грантів. Переможці отримають фінансування на проведення досліджень у розмірі 50 000 грн.",
     "date":"2024-11-11T16:00:00"},
    {"uid":"demo_006","sender":"admin@university.edu.ua",
     "subject":"Важливе повідомлення щодо нового пропускного режиму",
     "body":"Адміністрація університету повідомляє, що з 1 грудня вводиться новий порядок пропускного режиму. Для проходу необхідно пред'явити студентський квиток.",
     "date":"2024-11-10T11:00:00"},
    {"uid":"demo_007","sender":"erasmus@university.edu.ua",
     "subject":"Програма Erasmus+ обмін студентами 2025 рік",
     "body":"Запрошуємо студентів 2–4 курсів взяти участь у програмі академічного обміну Erasmus+ з університетами Польщі, Чехії та Австрії.",
     "date":"2024-11-09T13:30:00"},
    {"uid":"demo_008","sender":"council@university.edu.ua",
     "subject":"Прохання надати відгук щодо нового положення про навчання",
     "body":"Шановні колеги! Просимо надати ваші пропозиції та зауваження щодо проєкту нового Положення про організацію освітнього процесу.",
     "date":"2024-11-08T09:15:00"},
    {"uid":"demo_spam_01","sender":"newsletter@pinterest.com",
     "subject":"10 ідей для затишного інтер'єру — тільки для вас",
     "body":"Привіт! Ваша добірка від Pinterest готова. Переглядайте найкращі піни тижня.",
     "date":"2024-11-15T08:00:00"},
    {"uid":"demo_spam_02","sender":"noreply@rozetka.com.ua",
     "subject":"Ваше замовлення #8821044 відправлено",
     "body":"Трек-номер: 59000123456789. Очікуйте доставку від Nova Poshta.",
     "date":"2024-11-14T11:00:00"},
    {"uid":"demo_spam_03","sender":"info@facebookmail.com",
     "subject":"Ви маєте 3 нових сповіщення у Facebook",
     "body":"Іванко Петренко та ще 2 особи прокоментували ваше фото.",
     "date":"2024-11-13T09:30:00"},
    {"uid":"demo_spam_04","sender":"no-reply@monobank.ua",
     "subject":"Виписка по картці за жовтень 2024",
     "body":"Ваша виписка по картці за жовтень 2024 готова. Сума витрат: 3240 грн.",
     "date":"2024-11-12T07:00:00"},
    {"uid":"demo_spam_05","sender":"deals@aliexpress.com",
     "subject":"Знижки до 90% — тільки сьогодні",
     "body":"Не пропустіть найбільший розпродаж року на електроніку та аксесуари.",
     "date":"2024-11-11T10:00:00"},
    {"uid":"demo_spam_06","sender":"notifications@linkedin.com",
     "subject":"У вас 5 нових запитів на LinkedIn",
     "body":"Перегляньте хто хоче додати вас до своєї мережі.",
     "date":"2024-11-10T14:00:00"},
]

app = Flask(__name__)
app.secret_key = SECRET_KEY
app.config["SESSION_PERMANENT"] = False
app.jinja_env.filters["render_body"] = _render_body


# ── Auth decorator ───────────────────────────────────────────────
def login_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login", next=request.path))
        return f(*args, **kwargs)
    return decorated


# Мапи фонових зображень для кожної вкладки
_BG_MAP = {
    "/":            "bg_main.jpg",
    "/starred":     "bg_starred.jpg",
    "/stats":       "bg_stats.jpg",
    "/search":      "bg_search.jpg",
    "/spam-rules":  "bg_spam.jpg",
    "/settings":    "bg_settings.jpg",
    "/categories":  "bg_category.jpg",
    "/help":        "bg_settings.jpg",
}

# Унікальний фон для кожної категорії листів
_CATEGORY_BG = {
    "Актуальні новини":      "bg_main.jpg",
    "Наукова діяльність":    "bg_stats.jpg",
    "Освітня діяльність":    "bg_category.jpg",
    "Міжнародна діяльність": "bg_search.jpg",
    "Приймальна комісія":    "bg_settings.jpg",
    "До обговорення":        "bg_starred.jpg",
    "Вакансії":              "bg_spam.jpg",
    "Спам / Реклама":        "bg_spam.jpg",
    "Невизначено":           "bg_category.jpg",
}

def _sidebar_ctx():
    """Контекст для сайдбару (категорії + лічильники)."""
    stats  = get_stats()
    by_cat = stats["by_category"]
    all_cats = get_all_categories()
    # Вибираємо фон: спочатку точний збіг, потім prefix, потім категорія
    path = request.path
    bg = _BG_MAP.get(path)
    if not bg:
        for prefix, img in _BG_MAP.items():
            if len(prefix) > 1 and path.startswith(prefix):
                bg = img
                break
    if not bg:
        # Для /category/<name> — унікальний фон категорії
        if path.startswith("/category/"):
            cat_name = path[len("/category/"):]
            from urllib.parse import unquote
            cat_name = unquote(cat_name)
            bg = _CATEGORY_BG.get(cat_name, "bg_category.jpg")
        else:
            bg = "bg_category.jpg"
    return {
        "categories":     all_cats,
        "category_names": [c["name"] for c in all_cats],
        "unread":         {i["category"]: i["unread"] for i in by_cat},
        "count":          {i["category"]: i["cnt"]    for i in by_cat},
        "starred_count":  get_starred_count(),
        "bg_image":       bg,
    }


@app.before_request
def ensure_db():
    init_db()


# ── Auth ─────────────────────────────────────────────────────────
@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        u = request.form.get("username", "")
        p = request.form.get("password", "")
        if u == get_setting("auth_user", "admin") and \
           p == get_setting("auth_pass", "admin"):
            session["logged_in"] = True
            session["username"]  = u
            next_url = request.args.get("next") or url_for("index")
            return redirect(next_url)
        error = "Невірний логін або пароль"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ── Main pages ───────────────────────────────────────────────────
@app.route("/")
@login_required
def index():
    stats = get_stats()
    by_cat = stats["by_category"]
    return render_template(
        "index.html",
        total=stats["total"],
        **_sidebar_ctx(),
    )


@app.route("/category/<path:name>")
@login_required
def category(name):
    all_cat_names = get_category_names() + ["Спам / Реклама", "Невизначено"]
    if name not in all_cat_names:
        return "Категорія не знайдена", 404
    emails = get_emails_by_category(name)
    return render_template("category.html", category=name,
                           emails=emails, **_sidebar_ctx())


@app.route("/email/<int:email_id>")
@login_required
def email_detail(email_id):
    em = get_email_by_id(email_id)
    if not em:
        return "Лист не знайдено", 404
    mark_as_read(email_id)
    return render_template("email_detail.html", email=em, **_sidebar_ctx())


# ── Search ───────────────────────────────────────────────────────
@app.route("/search")
@login_required
def search():
    q   = request.args.get("q", "").strip()
    cat = request.args.get("category", "")
    results = []
    if q:
        results = search_emails(q, category=cat if cat else None)
    return render_template(
        "search.html", q=q, results=results,
        filter_category=cat, **_sidebar_ctx()
    )


# ── Export CSV ───────────────────────────────────────────────────
@app.route("/export")
@login_required
def export_excel():
    """Експорт листів у Excel (.xlsx) з форматуванням."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io as _io

    cat    = request.args.get("category", "")
    emails = get_all_emails_for_export(category=cat if cat else None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cat[:31] if cat else "Всі листи"

    # ── Кольори ──
    BLUE_FILL  = PatternFill("solid", fgColor="1F4E8C")
    ALT_FILL   = PatternFill("solid", fgColor="EEF3FB")
    SPAM_FILL  = PatternFill("solid", fgColor="FFF0F0")
    STAR_FILL  = PatternFill("solid", fgColor="FFFBEA")
    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Заголовки ──
    headers = ["#", "Відправник", "Тема", "Дата", "Категорія", "Впевненість", "Прочитано", "Важливий"]
    widths  = [5,   32,            48,     14,      24,           13,             11,           10]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        cell.fill      = BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22

    # ── Дані ──
    CONF_COLORS = {
        range(80, 101): "1A7A3C",
        range(60, 80):  "2E7D32",
        range(40, 60):  "E65100",
        range(0, 40):   "B71C1C",
    }

    for row_i, em in enumerate(emails, 2):
        is_spam    = em.get("category") == "Спам / Реклама"
        is_starred = bool(em.get("is_starred"))
        fill = SPAM_FILL if is_spam else (STAR_FILL if is_starred else (ALT_FILL if row_i % 2 == 0 else None))

        conf_pct = int((em.get("confidence") or 0) * 100)
        conf_color = "555555"
        for rng, color in CONF_COLORS.items():
            if conf_pct in rng:
                conf_color = color
                break

        values = [
            row_i - 1,
            em.get("sender", ""),
            em.get("subject", ""),
            (em.get("date") or "")[:16].replace("T", " "),
            em.get("category", ""),
            f"{conf_pct}%",
            "Так" if em.get("is_read") else "Ні",
            "⭐" if is_starred else "",
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.border    = border
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=(col == 3))
            if fill:
                cell.fill = fill
            if col == 5 and is_spam:
                cell.font = Font(color="CC0000", name="Calibri", size=10)
            elif col == 6:
                cell.font = Font(color=conf_color, bold=True, name="Calibri", size=10)
            else:
                cell.font = Font(name="Calibri", size=10)

        ws.row_dimensions[row_i].height = 16

    # ── Автофільтр ──
    ws.auto_filter.ref = f"A1:H{len(emails)+1}"
    ws.freeze_panes    = "A2"

    # ── Зведений аркуш ──
    ws2 = wb.create_sheet("Статистика")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14

    ws2.cell(1,1,"Категорія").font      = Font(bold=True, color="FFFFFF", name="Calibri")
    ws2.cell(1,1).fill                  = BLUE_FILL
    ws2.cell(1,1).alignment             = Alignment(horizontal="center")
    ws2.cell(1,2,"Кількість листів").font = Font(bold=True, color="FFFFFF", name="Calibri")
    ws2.cell(1,2).fill                  = BLUE_FILL
    ws2.cell(1,2).alignment             = Alignment(horizontal="center")
    ws2.cell(1,3,"Середня впевненість").font = Font(bold=True, color="FFFFFF", name="Calibri")
    ws2.cell(1,3).fill                  = BLUE_FILL
    ws2.cell(1,3).alignment             = Alignment(horizontal="center")

    from collections import defaultdict
    cat_stats = defaultdict(lambda: {"count": 0, "conf_sum": 0.0})
    for em in emails:
        c = em.get("category", "—")
        cat_stats[c]["count"]    += 1
        cat_stats[c]["conf_sum"] += float(em.get("confidence") or 0)

    for row_i, (c, st) in enumerate(sorted(cat_stats.items()), 2):
        avg_conf = st["conf_sum"] / st["count"] if st["count"] else 0
        ws2.cell(row_i, 1, c).font        = Font(name="Calibri", size=10)
        ws2.cell(row_i, 2, st["count"]).font = Font(name="Calibri", size=10)
        ws2.cell(row_i, 2).alignment      = Alignment(horizontal="center")
        ws2.cell(row_i, 3, f"{avg_conf*100:.1f}%").font = Font(name="Calibri", size=10)
        ws2.cell(row_i, 3).alignment      = Alignment(horizontal="center")
        if row_i % 2 == 0:
            for col in range(1, 4):
                ws2.cell(row_i, col).fill = ALT_FILL

    # ── Зберегти в буфер ──
    buf2 = _io.BytesIO()
    wb.save(buf2)
    buf2.seek(0)

    safe_cat = cat.replace(" ", "_").replace("/", "-") if cat else "all"
    filename = f"mailsort_{safe_cat}.xlsx"
    return Response(
        buf2.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


# ── Settings ─────────────────────────────────────────────────────
@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    keys = ["imap_host", "imap_port", "email_user", "imap_pass",
            "fetch_limit", "threshold", "sync_interval", "auth_user", "auth_pass"]

    if request.method == "POST":
        for k in keys:
            v = request.form.get(k, "").strip()
            if v:
                set_setting(k, v)
        # Перезапускаємо авто-синхронізацію з новим інтервалом
        new_interval = int(get_setting("sync_interval", "300"))
        if new_interval > 0:
            start_auto_sync(new_interval)
        else:
            stop_auto_sync()
        flash("Налаштування збережено ✓", "success")
        return redirect(url_for("settings"))

    current = {k: get_setting(k) for k in keys}
    return render_template("settings.html", settings=current, **_sidebar_ctx())


# ── Actions ──────────────────────────────────────────────────────
@app.route("/fetch", methods=["POST"])
@login_required
def fetch():
    try:
        from email_fetcher import fetch_and_classify
        s = fetch_and_classify()
        if s.get("error_msg"):
            flash(f"Помилка: {s['error_msg']}", "error")
        else:
            saved = s.get('saved', 0)
            msg = f"Додано {saved} нових листів" if saved > 0 else "Нових листів немає"
            flash(msg, "success")
    except Exception as e:
        flash(f"Помилка IMAP: {e}", "error")
    return redirect(url_for("index"))


@app.route("/demo", methods=["POST"])
@login_required
def demo():
    from database import save_email
    cnt = 0
    for em in DEMO_EMAILS:
        cat, conf = _lazy_classify(em["subject"], em["body"], sender=em["sender"])
        save_email(uid=em["uid"], sender=em["sender"],
                   subject=em["subject"], body=em["body"],
                   date=em["date"], category=cat, confidence=conf)
        cnt += 1
    flash(f"Демо: додано {cnt} листів!", "success")
    return redirect(url_for("index"))


@app.route("/reclassify/<int:email_id>", methods=["POST"])
@login_required
def reclassify(email_id):
    new_cat = request.form.get("category")
    all_valid = get_category_names() + ["Спам / Реклама", "Невизначено"]
    if new_cat in all_valid:
        # Зберігаємо корекцію до активного навчання
        # (тільки для «реальних» категорій, не для спаму — спам вчиться окремо)
        try:
            from database import get_email_by_id, add_user_correction
            em = get_email_by_id(email_id)
            if em and new_cat not in ("Спам / Реклама", "Невизначено") \
                  and em.get("category") != new_cat:
                add_user_correction(
                    email_id=email_id,
                    subject=em.get("subject", ""),
                    body=em.get("body", ""),
                    sender=em.get("sender", ""),
                    old_category=em.get("category", ""),
                    old_confidence=float(em.get("confidence") or 0.0),
                    new_category=new_cat,
                )
        except Exception as ex:
            print(f"[reclassify] Не вдалося зберегти корекцію: {ex}")

        reclassify_email(email_id, new_cat)
        flash(f"Переміщено до «{new_cat}»", "success")
        return redirect(url_for("category", name=new_cat))
    return redirect(url_for("email_detail", email_id=email_id))




@app.route("/email/<int:email_id>/spam-domain", methods=["POST"])
@login_required
def email_spam_domain(email_id):
    """Додає домен відправника у правила спаму і переміщує лист у Спам."""
    import re
    from database import get_email_by_id, reclassify_email
    email = get_email_by_id(email_id)
    if not email:
        flash("Лист не знайдено", "error")
        return redirect(url_for("index"))

    sender = email["sender"]
    # Витягуємо email-адресу з поля "Ім'я <email@domain>"
    match = re.search(r'<([^>]+)>', sender)
    raw_email = match.group(1).strip() if match else sender.strip()
    # Беремо домен
    domain = raw_email.split("@")[-1].lower().strip() if "@" in raw_email else raw_email.lower()

    if domain:
        ok = add_spam_rule("domain", domain)
        reclassify_email(email_id, "Спам / Реклама")
        if ok:
            flash(f"Домен «{domain}» додано у спам. Лист переміщено.", "success")
        else:
            reclassify_email(email_id, "Спам / Реклама")
            flash(f"Домен «{domain}» вже був у спамі. Лист переміщено.", "success")
    else:
        flash("Не вдалося визначити домен відправника", "error")

    return redirect(url_for("email_detail", email_id=email_id))


@app.route("/email/<int:email_id>/spam-sender", methods=["POST"])
@login_required
def email_spam_sender(email_id):
    """Додає точну адресу відправника у правила спаму."""
    import re
    from database import get_email_by_id, reclassify_email
    email_obj = get_email_by_id(email_id)
    if not email_obj:
        flash("Лист не знайдено", "error")
        return redirect(url_for("index"))

    sender = email_obj["sender"]
    match = re.search(r'<([^>]+)>', sender)
    raw_email = match.group(1).strip() if match else sender.strip()

    if "@" in raw_email:
        ok = add_spam_rule("domain", raw_email.lower())
        reclassify_email(email_id, "Спам / Реклама")
        if ok:
            flash(f"Адресу «{raw_email}» додано у спам. Лист переміщено.", "success")
        else:
            flash(f"Адреса вже була у спамі. Лист переміщено.", "success")
    else:
        flash("Не вдалося визначити адресу відправника", "error")

    return redirect(url_for("email_detail", email_id=email_id))



@app.route("/email/<int:email_id>/unblock-domain", methods=["POST"])
@login_required
def email_unblock_domain(email_id):
    """Видаляє домен відправника зі списку спаму."""
    import re
    from database import get_email_by_id, get_connection
    email_obj = get_email_by_id(email_id)
    if not email_obj:
        flash("Лист не знайдено", "error")
        return redirect(url_for("index"))

    sender = email_obj["sender"]
    match = re.search(r'<([^>]+)>', sender)
    raw_email = match.group(1).strip() if match else sender.strip()
    domain = raw_email.split("@")[-1].lower().strip() if "@" in raw_email else raw_email.lower()

    if domain:
        with get_connection() as conn:
            # Видаляємо і домен, і точну адресу
            conn.execute(
                "DELETE FROM spam_rules WHERE LOWER(value) = ? OR LOWER(value) = ?",
                (domain, raw_email.lower())
            )
        flash(f"Домен «{domain}» розблоковано.", "success")
    else:
        flash("Не вдалося визначити домен", "error")

    return redirect(url_for("email_detail", email_id=email_id))

# ── Stats page ───────────────────────────────────────────────────
@app.route("/stats")
@login_required
def stats():
    return render_template("stats.html", **_sidebar_ctx())


# ── API ──────────────────────────────────────────────────────────
@app.route("/api/stats")
@login_required
def api_stats():
    return jsonify(get_stats())


@app.route("/api/model-stats")
@login_required
def api_model_stats():
    from classifier import TRAINING_DATA, MODELS, CONFIDENCE_THRESHOLD
    from preprocessor import preprocess
    from sklearn.model_selection import train_test_split, StratifiedKFold, cross_val_score
    from sklearn.metrics import confusion_matrix, classification_report

    texts  = [preprocess(t, "") for t, _ in TRAINING_DATA]
    labels = [cat for _, cat in TRAINING_DATA]
    cats   = sorted(set(labels))

    cv_results = {}
    for name, builder in MODELS.items():
        scores = cross_val_score(
            builder(), texts, labels,
            cv=StratifiedKFold(n_splits=5, shuffle=True, random_state=42),
            scoring="f1_weighted",
        )
        cv_results[name] = {
            "mean":   round(float(scores.mean()), 4),
            "std":    round(float(scores.std()),  4),
            "scores": [round(float(s), 4) for s in scores],
        }

    best_name = max(cv_results, key=lambda k: cv_results[k]["mean"])
    X_tr, X_te, y_tr, y_te = train_test_split(
        texts, labels, test_size=0.2, random_state=42, stratify=labels
    )
    pipe = MODELS[best_name]()
    pipe.fit(X_tr, y_tr)
    y_pred  = pipe.predict(X_te)
    cm      = confusion_matrix(y_te, y_pred, labels=cats).tolist()
    report  = classification_report(y_te, y_pred, output_dict=True)

    return jsonify({
        "cv_results":       cv_results,
        "best_model":       best_name,
        "categories":       cats,
        "confusion_matrix": cm,
        "per_class_f1":   {c: round(report[c]["f1-score"],  3) for c in cats if c in report},
        "per_class_prec": {c: round(report[c]["precision"], 3) for c in cats if c in report},
        "per_class_rec":  {c: round(report[c]["recall"],    3) for c in cats if c in report},
        "overall_f1":     round(float(report["weighted avg"]["f1-score"]), 3),
        "threshold":      CONFIDENCE_THRESHOLD,
        "train_size":     len(TRAINING_DATA),
    })


@app.route("/api/classify", methods=["POST"])
def api_classify():
    data     = request.get_json(force=True)
    cat, conf = _lazy_classify(data.get("subject",""), data.get("body",""))
    return jsonify({"category": cat, "confidence": round(conf, 4)})



# ── Starred ───────────────────────────────────────────────────────
@app.route("/starred")
@login_required
def starred():
    emails = get_starred_emails()
    return render_template("starred.html", emails=emails, **_sidebar_ctx())


@app.route("/toggle-star/<int:email_id>", methods=["POST"])
@login_required
def toggle_star(email_id):
    new_state = toggle_starred(email_id)
    # AJAX або редірект
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"starred": new_state})
    next_url = request.form.get("next") or url_for("starred")
    return redirect(next_url)




@app.route("/api/fix-moodle")
@login_required
def api_fix_moodle():
    """
    Діагностика і примусове виправлення категорії Moodle.
    Шукає за sender І subject — не залежить від точного формату адреси.
    """
    from database import get_connection
    with get_connection() as conn:
        # 1. Показуємо що є "мудл-подібного" в БД
        samples = conn.execute("""
            SELECT id, sender, subject, category FROM emails
            WHERE LOWER(sender) LIKE '%moodle%'
               OR LOWER(subject) LIKE '%moodle%'
               OR LOWER(body)    LIKE '%moodle%'
            LIMIT 20
        """).fetchall()

        # 2. Виправляємо: sender або subject містить "moodle"
        r1 = conn.execute("""
            UPDATE emails SET category='Moodle', confidence=0.99
            WHERE (LOWER(sender)  LIKE '%moodle%'
                OR LOWER(subject) LIKE '%moodle%')
            AND category != 'Moodle'
            AND category != 'Спам / Реклама'
        """)

        # 3. Виправляємо: типові Moodle-теми навіть без слова "moodle"
        r2 = conn.execute("""
            UPDATE emails SET category='Moodle', confidence=0.95
            WHERE (
                LOWER(subject) LIKE '%дедлайн%'
                OR LOWER(subject) LIKE '%deadline%'
                OR LOWER(subject) LIKE '%нове завдання%'
                OR LOWER(subject) LIKE '%new submission%'
                OR LOWER(subject) LIKE '%нова оцінка%'
                OR LOWER(subject) LIKE '%форум%курс%'
            )
            AND LOWER(sender) LIKE '%chnu%'
            AND category != 'Moodle'
            AND category != 'Спам / Реклама'
        """)

        total_fixed = r1.rowcount + r2.rowcount

    return jsonify({
        "fixed":   total_fixed,
        "samples": [{"id": s["id"], "sender": s["sender"],
                     "subject": s["subject"][:60], "category": s["category"]}
                    for s in samples],
        "message": f"Виправлено {total_fixed} листів → Moodle"
    })


# ── Auto-sync API ─────────────────────────────────────────────────
@app.route("/api/sync-status")
@login_required
def sync_status():
    from database import get_setting
    interval = int(get_setting("sync_interval", "300"))
    return jsonify({
        "interval":  interval,
        "running":   _sync_thread is not None and _sync_thread.is_alive(),
        "last_sync": _last_sync_result,
    })


@app.route("/api/sync-now", methods=["POST"])
@login_required
def sync_now():
    """
    Запускає синхронізацію у фоновому потоці та негайно повертає відповідь.
    Клієнт опитує /api/sync-progress кожні ~500мс для відстеження прогресу.
    """
    global _manual_sync_thread
    if _manual_sync_thread is not None and _manual_sync_thread.is_alive():
        with _manual_sync_lock:
            return jsonify({"ok": True, "already_running": True,
                            "state": dict(_manual_sync_state)})
    _manual_sync_thread = threading.Thread(target=_run_manual_sync, daemon=True)
    _manual_sync_thread.start()
    return jsonify({"ok": True, "started": True})


@app.route("/api/sync-progress")
@login_required
def sync_progress():
    """Повертає поточний стан ручної синхронізації (для прогрес-бару)."""
    with _manual_sync_lock:
        snapshot = dict(_manual_sync_state)
    snapshot["running"] = (_manual_sync_thread is not None
                           and _manual_sync_thread.is_alive())
    return jsonify(snapshot)


@app.route("/api/set-sync-interval", methods=["POST"])
@login_required
def set_sync_interval():
    data = request.get_json(force=True)
    interval = int(data.get("interval", 300))
    set_setting("sync_interval", str(interval))
    if interval > 0:
        start_auto_sync(interval)
    else:
        stop_auto_sync()
    return jsonify({"ok": True, "interval": interval})


# ── Delete emails ─────────────────────────────────────────────────
@app.route("/delete-email/<int:email_id>", methods=["POST"])
@login_required
def delete_email_route(email_id):
    next_url = request.form.get("next", url_for("index"))
    delete_email(email_id)
    flash("Лист видалено", "success")
    return redirect(next_url)


@app.route("/delete-demo", methods=["POST"])
@login_required
def delete_demo_route():
    cnt = delete_demo_emails()
    flash(f"Видалено {cnt} демо-листів", "success")
    return redirect(url_for("index"))


@app.route("/delete-all", methods=["POST"])
@login_required
def delete_all_route():
    cnt = delete_all_emails()
    flash(f"Видалено всі {cnt} листів", "success")
    return redirect(url_for("index"))


# ── Spam rules management ─────────────────────────────────────────
@app.route("/spam-rules")
@login_required
def spam_rules():
    rules = get_spam_rules()
    return render_template("spam_rules.html", rules=rules, **_sidebar_ctx())


@app.route("/spam-rules/add", methods=["POST"])
@login_required
def add_spam_rule_route():
    rule_type = request.form.get("rule_type", "domain")
    value     = request.form.get("value", "").strip().lower()
    if value:
        ok = add_spam_rule(rule_type, value)
        flash(f"Додано: {value}" if ok else f"Вже існує: {value}",
              "success" if ok else "error")
    return redirect(url_for("spam_rules"))


@app.route("/spam-rules/delete/<int:rule_id>", methods=["POST"])
@login_required
def delete_spam_rule_route(rule_id):
    delete_spam_rule(rule_id)
    flash("Правило видалено", "success")
    return redirect(url_for("spam_rules"))


# ── User categories management ────────────────────────────────────
@app.route("/categories")
@login_required
def categories_page():
    all_cats = get_all_categories()
    return render_template("categories.html", all_cats=all_cats, **_sidebar_ctx())


@app.route("/categories/add", methods=["POST"])
@login_required
def add_category_route():
    name = request.form.get("name", "").strip()
    icon = request.form.get("icon", "📂").strip() or "📂"
    if name:
        ok = add_category(name, icon)
        flash(f"Категорію «{name}» додано!" if ok
              else f"Категорія «{name}» вже існує", "success" if ok else "error")
    return redirect(url_for("categories_page"))


@app.route("/categories/delete", methods=["POST"])
@login_required
def delete_category_route():
    name = request.form.get("name", "")
    ok, msg = delete_category(name)
    flash(msg if not ok else f"Категорію «{name}» видалено", "success" if ok else "error")
    return redirect(url_for("categories_page"))

@app.route("/help")
@login_required
def help_page():
    return render_template("help.html", **_sidebar_ctx())



def debug_template():
    import os
    path = os.path.abspath(app.template_folder)
    files = os.listdir(path) if os.path.exists(path) else []
    base_path = os.path.join(path, "base.html")
    base_size = os.path.getsize(base_path) if os.path.exists(base_path) else 0
    return f"<pre>Template folder: {path}\nFiles: {files}\nbase.html size: {base_size} bytes</pre>"


@app.route("/api/chnu-news-save", methods=["POST"])
@login_required
def api_chnu_news_save():
    """Зберігає новини отримані браузером в кеш."""
    try:
        from news_fetcher import _ensure_table, _save
        _ensure_table()
        data = request.get_json(force=True)
        news = data.get("news", [])
        if news:
            _save(news)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


_training_status = {"running": False, "log": [], "done": False, "error": "", "score": 0, "total": 0, "current_cat": "", "cats_done": 0, "cats_total": 7}

# ── ChNU Training ─────────────────────────────────────────────────
@app.route("/api/train-chnu", methods=["POST"])
@login_required
def api_train_chnu():
    """Запускає навчання на новинах ЧНУ у фоновому потоці."""
    import threading
    
    def _do_train():
        try:
            import sys, os
            sys.path.insert(0, os.path.dirname(__file__))
            from chnu_trainer import CATEGORY_MAP, scrape_category, save_data, retrain, BASE, EXTRA_DATA_PATH
            import json, time

            _training_status["running"] = True
            _training_status["log"] = ["Починаємо завантаження новин ЧНУ..."]
            _training_status["done"] = False
            _training_status["error"] = ""
            _training_status["cats_done"] = 0
            _training_status["cats_total"] = len(CATEGORY_MAP)

            all_examples = []
            for idx, (category, urls) in enumerate(CATEGORY_MAP.items()):
                _training_status["current_cat"] = category
                _training_status["log"].append(f"[{category}] завантаження...")
                examples = scrape_category(
                    category,
                    [BASE + u for u in urls],
                    verbose=False, max_articles=15
                )
                count = len(examples)
                _training_status["cats_done"] = idx + 1
                if count > 0:
                    _training_status["log"].append(f"[{category}] зібрано {count} прикладів ✓")
                else:
                    _training_status["log"].append(f"[{category}] нічого не знайдено — сторінка недоступна")
                all_examples.extend(examples)
                time.sleep(0.3)

            if not all_examples:
                _training_status["error"] = "Нічого не знайдено. Перевірте інтернет-з'єднання."
                _training_status["running"] = False
                return

            save_data(all_examples)
            _training_status["log"].append(f"Всього {len(all_examples)} прикладів. Перенавчання моделі...")

            score = retrain(all_examples, verbose=False)

            # Скидаємо кеш моделі
            import classifier
            classifier._pipeline = None

            _training_status["log"].append(f"Готово! F1 = {score:.1%}")
            _training_status["score"] = round(score, 4)
            _training_status["total"] = len(all_examples)
            _training_status["done"] = True

        except Exception as e:
            _training_status["error"] = str(e)
        finally:
            _training_status["running"] = False

    if _training_status.get("running"):
        return jsonify({"ok": False, "error": "Навчання вже виконується"})

    t = threading.Thread(target=_do_train, daemon=True)
    t.start()
    return jsonify({"ok": True, "message": "Навчання розпочато"})


@app.route("/api/train-status")
@login_required
def api_train_status():
    return jsonify(_training_status)



@app.route("/api/news-debug")
@login_required
def api_news_debug():
    """Діагностика новин — показує що саме не працює."""
    try:
        from news_fetcher import debug_fetch
        results = debug_fetch()
        return jsonify({"ok": True, "results": results})
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()})

# ── ChNU News ────────────────────────────────────────────────────
@app.route("/api/chnu-news")
@login_required
def api_chnu_news():
    try:
        from news_fetcher import fetch_news
        force = request.args.get("force") == "1"
        news  = fetch_news(force=force)
        return jsonify({"ok": True, "news": news[:15]})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "news": []})




@app.route("/api/chnu-article")
@login_required
def api_chnu_article():
    """Завантажує текст статті ЧНУ за URL."""
    url = request.args.get("url", "").strip()
    if not url or not url.startswith("http"):
        return jsonify({"ok": False, "error": "Невірний URL"})
    try:
        from chnu_trainer import get_article_text
        import requests as _req
        from bs4 import BeautifulSoup as _BS
        # Завантажуємо повний HTML для отримання заголовку і тексту
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept-Language": "uk-UA,uk;q=0.9",
        }
        r = _req.get(url, timeout=10, headers=headers)
        r.raise_for_status()
        soup = _BS(r.text, "html.parser")
        # Заголовок
        title = ""
        og_title = soup.find("meta", property="og:title")
        if og_title:
            title = og_title.get("content", "")
        if not title:
            h1 = soup.find("h1")
            title = h1.get_text(strip=True) if h1 else ""
        # Зображення
        image = ""
        og_img = soup.find("meta", property="og:image")
        if og_img:
            image = og_img.get("content", "")
        # Дата
        date = ""
        time_el = soup.find("time")
        if time_el:
            date = time_el.get("datetime", "") or time_el.get_text(strip=True)
        # Текст
        text = get_article_text(url)
        return jsonify({"ok": True, "title": title, "text": text, "image": image, "date": date[:10], "url": url})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ═══ ML-якість та активне навчання ═════════════════════════════════

_retrain_state = {
    "phase":    "idle",   # idle | running | done | error
    "message":  "",
    "started_at":  None,
    "finished_at": None,
    "result": None,
}
_retrain_lock   = threading.Lock()
_retrain_thread = None


def _run_retrain():
    """Тіло фонового потоку перенавчання."""
    try:
        with _retrain_lock:
            _retrain_state.update({
                "phase": "running",
                "message": "Збираю корпус…",
                "started_at":  time.time(),
                "finished_at": None,
                "result": None,
            })

        # lazy import, бо classifier тягне sklearn
        from classifier import train, _load_corpus
        texts, labels, sources, n_corr = _load_corpus()

        with _retrain_lock:
            _retrain_state["message"] = (
                f"Навчання на {len(texts)} прикладах "
                f"({len(texts) - n_corr*2} базових + {n_corr} корекцій)…"
            )

        train(print_report=False)

        # Беремо останню записану версію
        from database import get_model_versions, get_corrections_stats
        versions = get_model_versions(limit=1)
        stats    = get_corrections_stats()
        latest   = versions[0] if versions else {}

        with _retrain_lock:
            _retrain_state.update({
                "phase":   "done",
                "message": (f"Готово: {latest.get('model_name','?')} "
                            f"F1={latest.get('f1_weighted',0):.3f}"),
                "finished_at": time.time(),
                "result": {
                    "model_name":  latest.get("model_name"),
                    "n_samples":   latest.get("n_samples"),
                    "n_corrections": latest.get("n_corrections"),
                    "f1_weighted": latest.get("f1_weighted"),
                    "accuracy":    latest.get("accuracy"),
                    "pending_after": stats.get("pending", 0),
                },
            })
    except Exception as e:
        import traceback; traceback.print_exc()
        with _retrain_lock:
            _retrain_state.update({
                "phase": "error",
                "message": str(e),
                "finished_at": time.time(),
            })


@app.route("/ml-quality")
@login_required
def ml_quality():
    """Сторінка діагностики ML-моделі: метрики + матриця помилок + історія."""
    from database import get_corrections_stats, get_model_versions
    from classifier import diagnose, get_threshold, get_keyword_markers
    try:
        diag = diagnose()
    except Exception as e:
        import traceback; traceback.print_exc()
        diag = {"error": str(e), "labels": [], "matrix": [],
                "per_class": {}, "accuracy": 0, "f1_weighted": 0,
                "top_mistakes": [], "test_size": 0, "train_size": 0}

    try:
        markers = get_keyword_markers()
    except Exception:
        markers = {}

    return render_template(
        "ml_quality.html",
        diag=diag,
        threshold=get_threshold(),
        corr_stats=get_corrections_stats(),
        versions=get_model_versions(limit=10),
        markers=markers,
        **_sidebar_ctx(),
    )


@app.route("/api/retrain", methods=["POST"])
@login_required
def api_retrain():
    """Запускає перенавчання моделі у фоновому потоці."""
    global _retrain_thread
    if _retrain_thread is not None and _retrain_thread.is_alive():
        with _retrain_lock:
            return jsonify({"ok": True, "already_running": True,
                            "state": dict(_retrain_state)})
    _retrain_thread = threading.Thread(target=_run_retrain, daemon=True)
    _retrain_thread.start()
    return jsonify({"ok": True, "started": True})


@app.route("/api/retrain-progress")
@login_required
def api_retrain_progress():
    """Стан перенавчання для UI."""
    with _retrain_lock:
        snapshot = dict(_retrain_state)
    snapshot["running"] = (_retrain_thread is not None
                           and _retrain_thread.is_alive())
    return jsonify(snapshot)


@app.route("/api/corrections-stats")
@login_required
def api_corrections_stats():
    """Для лічильника в бічній панелі: скільки корекцій чекають перенавчання."""
    from database import get_corrections_stats
    return jsonify(get_corrections_stats())


# ── Ініціалізація (працює і з gunicorn, і з python app.py) ────────
init_db()
from database import get_setting as _gs
_interval = int(_gs('sync_interval', '300'))
if _interval > 0:
    start_auto_sync(_interval)

if __name__ == "__main__":
    print()
    print("  MailSort запущено!")
    print("  ->  http://localhost:5000")
    print()
    _port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, port=_port, use_reloader=False)
